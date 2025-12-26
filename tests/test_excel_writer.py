import os
from datetime import date

import pytest
from openpyxl import Workbook, load_workbook

# リファクタリング後のモジュールをインポート
from app.services.excel import writer

# マッピング定義も参照して、正しいセルに書き込まれているか確認する
from app.services.excel.mappings import SELECTION_MAPPING, TEXT_MAPPING


@pytest.fixture
def temp_excel_template(tmp_path):
    """
    テスト用にダミーのExcelテンプレートを作成するフィクスチャ。
    mappings.py で使用されているシート名が含まれている必要があります。
    """
    wb = Workbook()
    # デフォルトのシート名を変更
    ws1 = wb.active
    ws1.title = "様式23_1"
    # もう一つのシートを作成
    wb.create_sheet("様式23_2")

    # 一時ファイルとして保存
    template_path = tmp_path / "test_template.xlsx"
    wb.save(template_path)
    return str(template_path)

@pytest.fixture
def output_dir(tmp_path):
    """テスト用の出力ディレクトリ"""
    d = tmp_path / "output"
    d.mkdir()
    return str(d)

def test_excel_write_basic_text(temp_excel_template, output_dir, monkeypatch):
    """
    【正常系】基本的なテキストデータが狙ったセル（F3等）に書き込まれるかテスト
    """
    # 1. 設定のパッチ: テスト用のテンプレートと出力先を強制する
    monkeypatch.setattr(writer, "TEMPLATE_PATH", temp_excel_template)
    monkeypatch.setattr(writer, "OUTPUT_DIR", output_dir)

    # 2. テストデータ準備 (mappings.py の name は "様式23_1", "F3" と定義されている前提)
    input_data = {
        "name": "テスト患者",
        "age": 88
    }

    # 3. 実行
    output_file = writer.create_plan_sheet(input_data)

    # 4. 検証: 生成されたファイルを開いて中身を確認
    assert os.path.exists(output_file)

    wb = load_workbook(output_file)
    ws = wb["様式23_1"]

    # マッピング定義通りの場所に書き込まれているか
    name_sheet, name_cell = TEXT_MAPPING["name"]
    assert ws[name_cell].value == "テスト患者"

    # 年齢の確認
    age_sheet, age_cell = TEXT_MAPPING["age"]
    assert wb[age_sheet][age_cell].value == 88

def test_excel_write_selection_radio(temp_excel_template, output_dir, monkeypatch):
    """
    【ロジック】ラジオボタン（選択肢）のロジックが機能し、チェックマークが入るかテスト
    """
    monkeypatch.setattr(writer, "TEMPLATE_PATH", temp_excel_template)
    monkeypatch.setattr(writer, "OUTPUT_DIR", output_dir)

    # 住宅種別: 'home_detached' (戸建て) を選択
    # mappings.pyによれば、E3 と H3 にチェックが入るはず
    input_data = {
        "goal_p_residence_slct": "home_detached"
    }

    output_file = writer.create_plan_sheet(input_data)
    wb = load_workbook(output_file)

    # SELECTION_MAPPING から期待されるセル座標を取得して検証
    target_cells = SELECTION_MAPPING["goal_p_residence_slct"]["home_detached"]

    for sheet_name, cell_addr in target_cells:
        assert wb[sheet_name][cell_addr].value == "☑", f"{cell_addr} にチェックが入っていません"

    # 選択されていないセル（例: マンションのK3）は空またはチェックなしであること
    # (初期状態によるが、今回のロジックでは明示的に "☐" を入れるかは実装依存。
    #  create_plan_sheetの実装ではデフォルト値の上書きなので、元のテンプレート次第)

def test_excel_write_dates(temp_excel_template, output_dir, monkeypatch):
    """
    【ロジック】日付オブジェクトが 年・月・日 のセルに分割されて書き込まれるかテスト
    """
    monkeypatch.setattr(writer, "TEMPLATE_PATH", temp_excel_template)
    monkeypatch.setattr(writer, "OUTPUT_DIR", output_dir)

    test_date = date(2025, 12, 25)
    input_data = {
        "header_evaluation_date": test_date
    }

    output_file = writer.create_plan_sheet(input_data)
    wb = load_workbook(output_file)
    ws1 = wb["様式23_1"]

    # DATE_MAPPING の定義: "header_evaluation": ("様式23_1", "AN3", "AQ3", "AT3")
    assert ws1["AN3"].value == 2025
    assert ws1["AQ3"].value == 12
    assert ws1["AT3"].value == 25

def test_create_plan_sheet_returns_bytes(temp_excel_template, output_dir, monkeypatch):
    """
    【プレビュー用】ファイル保存ではなく、バイト列(BytesIO)が返るモードのテスト
    """
    monkeypatch.setattr(writer, "TEMPLATE_PATH", temp_excel_template)
    monkeypatch.setattr(writer, "OUTPUT_DIR", output_dir)

    input_data = {"name": "Preview Test"}

    # return_bytes=True を指定
    result_bytes = writer.create_plan_sheet(input_data, return_bytes=True)

    # BytesIOオブジェクトであることを確認
    assert hasattr(result_bytes, "read")
    assert hasattr(result_bytes, "seek")

    # 実際にExcelとして開けるか確認
    result_bytes.seek(0)
    wb = load_workbook(result_bytes)
    assert wb["様式23_1"]["F3"].value == "Preview Test"
