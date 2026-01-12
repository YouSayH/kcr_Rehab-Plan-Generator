# app/services/excel/writer.py
import os
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Font

# 設定ファイルからマッピングを読み込み
from .mappings import DATE_MAPPING, GENDER_MAPPING, OUTPUT_DIR, SELECTION_MAPPING, TEMPLATE_PATH, TEXT_MAPPING


def _get_cell_by_address(wb, sheet_name, cell_address):
    """シート名とセル座標からセルオブジェクトを取得する（結合セル対応）"""
    try:
        ws = wb[sheet_name]
        cell = ws[cell_address]
        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cell_ranges:
                if cell.coordinate in merged_range:
                    return merged_range.min_cell
        return cell
    except Exception as e:
        print(f"   [エラー] シート '{sheet_name}' またはセル '{cell_address}' の取得に失敗: {e}")
        return None

def _write_date_fields(wb, plan_data):
    """日付項目の書き込み処理"""
    for key_prefix, (sheet_name, year_addr, month_addr, day_addr) in DATE_MAPPING.items():
        # データ内には "header_evaluation_date" のようなキーで入っていると想定
        date_key = f"{key_prefix}_date"
        date_value = plan_data.get(date_key)

        if isinstance(date_value, date):
            try:
                _get_cell_by_address(wb, sheet_name, year_addr).value = date_value.year
                _get_cell_by_address(wb, sheet_name, month_addr).value = date_value.month
                _get_cell_by_address(wb, sheet_name, day_addr).value = date_value.day
                print(f"   [成功] 日付書き込み: {date_key} -> {date_value}")
            except Exception as e:
                print(f"   [エラー] 日付 '{date_key}' の書き込み中にエラー: {e}")

def _write_gender_field(wb, plan_data):
    """性別項目の書き込み処理"""
    gender = plan_data.get("gender")
    if not gender:
        return

    m_conf = GENDER_MAPPING
    male_cell = _get_cell_by_address(wb, *m_conf["male_cell"])
    female_cell = _get_cell_by_address(wb, *m_conf["female_cell"])

    if not (male_cell and female_cell):
        return

    font_selected = Font(size=13, bold=False)
    font_unselected = Font(size=11, bold=False)

    if gender == m_conf["male_value"]:
        male_cell.value, male_cell.font = m_conf["selected_mark"], font_selected
        female_cell.value, female_cell.font = m_conf["unselected_mark_f"], font_unselected
    elif gender == m_conf["female_value"]:
        male_cell.value, male_cell.font = m_conf["unselected_mark_m"], font_unselected
        female_cell.value, female_cell.font = m_conf["selected_mark_f"], font_selected
    else:
        male_cell.value, male_cell.font = m_conf["unselected_mark_m"], font_unselected
        female_cell.value, female_cell.font = m_conf["unselected_mark_f"], font_unselected

def _write_selection_fields(wb, plan_data):
    """選択肢（ラジオボタン等）の書き込み処理"""
    for db_key, mapping_rules in SELECTION_MAPPING.items():
        user_value = plan_data.get(db_key)
        if not user_value:
            continue

        target_rule = mapping_rules.get(user_value)
        if target_rule:
            for rule in target_rule:
                # rule は (sheet, address) または (sheet, address, write_value)
                sheet_name = rule[0]
                address = rule[1]
                write_val = rule[2] if len(rule) > 2 else "☑"

                cell = _get_cell_by_address(wb, sheet_name, address)
                if cell:
                    cell.value = write_val
                    print(f"   [情報] 選択肢処理: {db_key}={user_value} -> {address}={write_val}")

def create_plan_sheet(plan_data, return_bytes=False):
    """【リファクタリング版】Excelに計画書を書き込む"""
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

    try:
        wb = load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        print(f"[エラー] テンプレートファイルが見つかりません: {TEMPLATE_PATH}")
        raise

    # 1. 基本テキスト・チェックボックスの書き込み
    for db_col_name, (sheet_name, cell_address) in TEXT_MAPPING.items():
        # 日付や性別など特殊処理に含まれるキーがTEXT_MAPPINGに残っている場合のガード
        if "date" in db_col_name or db_col_name == "gender":
            continue

        value = plan_data.get(db_col_name)
        if value is None or value == "":
            continue

        target_cell = _get_cell_by_address(wb, sheet_name, cell_address)
        if target_cell:
            try:
                # ブール値または `_chk` キーはチェックマークに変換
                if isinstance(value, bool) or db_col_name.endswith("_chk"):
                    target_cell.value = "☑" if value else "☐"
                else:
                    target_cell.value = value
            except Exception as e:
                print(f"   [エラー] 書き込み失敗 {db_col_name}: {e}")

    # 2. 特殊処理の実行
    _write_date_fields(wb, plan_data)
    _write_gender_field(wb, plan_data)
    _write_selection_fields(wb, plan_data)

    # 3. ファイルの保存
    if return_bytes:
        import io
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_patient_name = "".join(c for c in plan_data.get("name", "NoName") if c.isalnum())
    output_filename = f"RehabPlan_{safe_patient_name}_{timestamp}.xlsx"
    output_filepath = os.path.join(OUTPUT_DIR, output_filename)

    wb.save(output_filepath)
    print(f"\n計画書を {output_filepath} に保存しました。")

    return output_filepath
